"""
Jeopardy Translation & Adaptation Engine
Translates Jeopardy! questions to Georgian, analyzes linguistic adaptability,
caches repeated translations, and exports to formatted Excel (.xlsx) and CSV.
Supports continuous full-dataset background execution with checkpointing.
"""

import os
import sys
import re
import html
import time
import sqlite3
import csv
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from deep_translator import GoogleTranslator
from linguistic_filter import analyze_adaptability

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def clean_html(raw_html: str) -> str:
    """Removes HTML tags and decodes HTML entities."""
    if not raw_html:
        return ""
    cleanr = re.compile(r'<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return html.unescape(cleantext).strip()


class JeopardyTranslatorEngine:
    def __init__(self, db_path: str = "jeopardy_translated.db", num_threads: int = 12):
        self.db_path = db_path
        self.num_threads = num_threads
        self.lock = Lock()
        self.category_cache = {}
        self.init_db()
        self.load_category_cache()

    def init_db(self):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS translations (
                    row_id INTEGER PRIMARY KEY,
                    show_number TEXT,
                    air_date TEXT,
                    round TEXT,
                    category_en TEXT,
                    category_ka TEXT,
                    value TEXT,
                    question_en TEXT,
                    question_ka TEXT,
                    answer_en TEXT,
                    answer_ka TEXT,
                    is_adaptable TEXT,
                    adaptability_reason TEXT,
                    status TEXT DEFAULT 'PENDING'
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS category_cache (
                    category_en TEXT PRIMARY KEY,
                    category_ka TEXT
                )
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_status ON translations(status)")
            conn.commit()

    def load_category_cache(self):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            for cat_en, cat_ka in cursor.execute("SELECT category_en, category_ka FROM category_cache"):
                self.category_cache[cat_en] = cat_ka
        print(f"[INFO] Loaded {len(self.category_cache):,} cached categories.", flush=True)

    def save_category_translation(self, cat_en: str, cat_ka: str):
        with self.lock:
            self.category_cache[cat_en] = cat_ka
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "INSERT OR REPLACE INTO category_cache (category_en, category_ka) VALUES (?, ?)",
                    (cat_en, cat_ka)
                )
                conn.commit()

    def import_csv_if_needed(self, csv_path: str = "JEOPARDY_CSV.csv"):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM translations")
            count = cursor.fetchone()[0]
            if count > 0:
                print(f"[INFO] Database already contains {count:,} records.", flush=True)
                return count

        print(f"[INFO] Importing {csv_path} into SQLite database...", flush=True)
        records = []
        with open(csv_path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            header = [c.strip() for c in next(reader)]
            
            for idx, row in enumerate(reader, start=1):
                if len(row) < 7:
                    continue
                show_num = row[0].strip()
                air_date = row[1].strip()
                round_name = row[2].strip()
                category = clean_html(row[3].strip())
                value = row[4].strip()
                question = clean_html(row[5].strip())
                answer = clean_html(row[6].strip())

                is_adaptable, reason = analyze_adaptability(category, question, answer)

                records.append((
                    idx, show_num, air_date, round_name,
                    category, "", value,
                    question, "", answer, "",
                    is_adaptable, reason, 'PENDING'
                ))

                if len(records) >= 10000:
                    with sqlite3.connect(self.db_path) as conn:
                        conn.executemany("""
                            INSERT OR IGNORE INTO translations (
                                row_id, show_number, air_date, round,
                                category_en, category_ka, value,
                                question_en, question_ka, answer_en, answer_ka,
                                is_adaptable, adaptability_reason, status
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, records)
                        conn.commit()
                    records = []

            if records:
                with sqlite3.connect(self.db_path) as conn:
                    conn.executemany("""
                        INSERT OR IGNORE INTO translations (
                            row_id, show_number, air_date, round,
                            category_en, category_ka, value,
                            question_en, question_ka, answer_en, answer_ka,
                            is_adaptable, adaptability_reason, status
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, records)
                    conn.commit()

        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM translations")
            total = cursor.fetchone()[0]
        print(f"[SUCCESS] Imported {total:,} records.", flush=True)
        return total

    def translate_single_item(self, row, translator: GoogleTranslator):
        row_id, show_num, air_date, round_name, cat_en, val, q_en, a_en, is_adapt, adapt_reason = row
        
        # Translate category (cached)
        cat_ka = self.category_cache.get(cat_en)
        if not cat_ka and cat_en:
            try:
                cat_ka = translator.translate(cat_en)
                self.save_category_translation(cat_en, cat_ka)
            except Exception:
                cat_ka = cat_en

        # Translate Question
        q_ka = ""
        if q_en:
            retries = 3
            for attempt in range(retries):
                try:
                    q_ka = translator.translate(q_en)
                    break
                except Exception:
                    if attempt == retries - 1:
                        q_ka = q_en
                    time.sleep(1 + attempt * 2)

        # Translate Answer
        a_ka = ""
        if a_en:
            retries = 3
            for attempt in range(retries):
                try:
                    a_ka = translator.translate(a_en)
                    break
                except Exception:
                    if attempt == retries - 1:
                        a_ka = a_en
                    time.sleep(1 + attempt * 2)

        return (row_id, cat_ka, q_ka, a_ka)

    def process_batch(self, limit: int = None, chunk_size: int = 500, auto_export_interval: int = 2000):
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM translations WHERE status = 'PENDING'")
            total_pending = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM translations WHERE status = 'DONE'")
            total_done = cursor.fetchone()[0]

        if total_pending == 0:
            print("[INFO] All questions have already been translated!", flush=True)
            return 0

        target_count = min(limit, total_pending) if limit else total_pending
        print(f"[START] Translating {target_count:,} questions (Already translated: {total_done:,} / Total: {total_pending + total_done:,}) with {self.num_threads} threads...", flush=True)

        overall_processed = 0
        since_last_export = 0
        start_time = time.time()

        while overall_processed < target_count:
            fetch_limit = min(chunk_size, target_count - overall_processed)
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute(f"""
                    SELECT row_id, show_number, air_date, round, category_en, value, question_en, answer_en, is_adaptable, adaptability_reason
                    FROM translations
                    WHERE status = 'PENDING'
                    ORDER BY row_id ASC
                    LIMIT {fetch_limit}
                """)
                pending_rows = cursor.fetchall()

            if not pending_rows:
                break

            batch_updates = []
            def worker_task(item):
                translator = GoogleTranslator(source='en', target='ka')
                return self.translate_single_item(item, translator)

            with ThreadPoolExecutor(max_workers=self.num_threads) as executor:
                future_to_id = {executor.submit(worker_task, row): row[0] for row in pending_rows}

                for future in as_completed(future_to_id):
                    try:
                        res = future.result()
                        batch_updates.append((res[1], res[2], res[3], 'DONE', res[0]))
                        overall_processed += 1
                        since_last_export += 1

                        if len(batch_updates) >= 50:
                            with sqlite3.connect(self.db_path) as conn:
                                cursor = conn.cursor()
                                cursor.executemany("""
                                    UPDATE translations
                                    SET category_ka = ?, question_ka = ?, answer_ka = ?, status = ?
                                    WHERE row_id = ?
                                """, batch_updates)
                                conn.commit()
                            batch_updates = []

                        if overall_processed % 100 == 0 or overall_processed == target_count:
                            elapsed = time.time() - start_time
                            rate = overall_processed / elapsed if elapsed > 0 else 0
                            pct = (total_done + overall_processed) / (total_done + total_pending) * 100
                            print(f"[PROGRESS] Done: {total_done + overall_processed:,}/{total_done + total_pending:,} ({pct:.2f}%) - Speed: {rate:.1f} q/s - ETA: {(target_count - overall_processed)/rate/60:.1f} min", flush=True)

                    except Exception as exc:
                        r_id = future_to_id[future]
                        print(f"[ERROR] Row {r_id}: {exc}", flush=True)

            if batch_updates:
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    cursor.executemany("""
                        UPDATE translations
                        SET category_ka = ?, question_ka = ?, answer_ka = ?, status = ?
                        WHERE row_id = ?
                    """, batch_updates)
                    conn.commit()

            if since_last_export >= auto_export_interval:
                print(f"[AUTO-SAVE] Exporting current progress ({total_done + overall_processed:,} rows) to Excel and CSV...", flush=True)
                self.export_to_csv("JEOPARDY_GEORGIAN.csv")
                self.export_to_excel("JEOPARDY_GEORGIAN.xlsx")
                since_last_export = 0

        # Final export
        print(f"[FINISH] Exporting final results to Excel and CSV...", flush=True)
        self.export_to_csv("JEOPARDY_GEORGIAN.csv")
        self.export_to_excel("JEOPARDY_GEORGIAN.xlsx")
        print(f"[ALL DONE] Processed {overall_processed:,} items in {(time.time() - start_time)/60:.2f} minutes.", flush=True)
        return overall_processed

    def export_to_excel(self, output_file: str = "JEOPARDY_GEORGIAN.xlsx", limit: int = None):
        if not OPENPYXL_AVAILABLE:
            print("[WARNING] openpyxl not installed. Exporting to CSV instead.", flush=True)
            self.export_to_csv(output_file.replace(".xlsx", ".csv"), limit=limit)
            return

        print(f"[INFO] Exporting to Excel: {output_file} ...", flush=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jeopardy Georgian"

        headers = [
            "ID", "Show Number", "Air Date", "Round",
            "Category (EN)", "Category (KA)", "Value",
            "Question (EN)", "Question (KA)",
            "Answer (EN)", "Answer (KA)",
            "მოერგება ქართულად?", "თავსებადობის მიზეზი"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fill_yes = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # green
        fill_no = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # red
        fill_part = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # yellow

        font_bold = Font(name="Calibri", size=10, bold=True)

        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            query = """
                SELECT row_id, show_number, air_date, round,
                       category_en, category_ka, value,
                       question_en, question_ka,
                       answer_en, answer_ka,
                       is_adaptable, adaptability_reason
                FROM translations
                WHERE status = 'DONE'
                ORDER BY row_id ASC
            """
            if limit:
                query += f" LIMIT {limit}"
            cursor.execute(query)
            
            row_count = 1
            for row in cursor:
                row_count += 1
                ws.append(list(row))
                
                adapt_cell = ws.cell(row=row_count, column=12)
                adapt_cell.font = font_bold
                if row[11] == "კი":
                    adapt_cell.fill = fill_yes
                    adapt_cell.alignment = Alignment(horizontal="center")
                elif row[11] == "არა":
                    adapt_cell.fill = fill_no
                    adapt_cell.alignment = Alignment(horizontal="center")
                else:
                    adapt_cell.fill = fill_part
                    adapt_cell.alignment = Alignment(horizontal="center")

        col_widths = {
            "A": 8, "B": 12, "C": 12, "D": 16,
            "E": 25, "F": 25, "G": 10,
            "H": 45, "I": 45,
            "J": 20, "K": 20,
            "L": 18, "M": 40
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        ws.freeze_panes = "A2"
        wb.save(output_file)
        print(f"[SUCCESS] Successfully saved {row_count - 1:,} rows to {output_file}", flush=True)

    def export_to_csv(self, output_file: str = "JEOPARDY_GEORGIAN.csv", limit: int = None):
        print(f"[INFO] Exporting to CSV (UTF-8 BOM): {output_file} ...", flush=True)
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.cursor()
            query = """
                SELECT row_id, show_number, air_date, round,
                       category_en, category_ka, value,
                       question_en, question_ka,
                       answer_en, answer_ka,
                       is_adaptable, adaptability_reason
                FROM translations
                WHERE status = 'DONE'
                ORDER BY row_id ASC
            """
            if limit:
                query += f" LIMIT {limit}"
            cursor.execute(query)
            
            with open(output_file, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "ID", "Show Number", "Air Date", "Round",
                    "Category (EN)", "Category (KA)", "Value",
                    "Question (EN)", "Question (KA)",
                    "Answer (EN)", "Answer (KA)",
                    "Is Adaptable (KA)", "Adaptability Reason"
                ])
                count = 0
                for row in cursor:
                    writer.writerow(row)
                    count += 1
        print(f"[SUCCESS] Successfully saved {count:,} rows to {output_file}", flush=True)


def main():
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

    parser = argparse.ArgumentParser(description="Jeopardy Georgian Translation Engine")
    parser.add_argument("--import-csv", action="store_true", help="Import JEOPARDY_CSV.csv into database")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of rows to process")
    parser.add_argument("--threads", type=int, default=12, help="Number of worker threads")
    parser.add_argument("--chunk-size", type=int, default=500, help="Chunk size for batch processing")
    parser.add_argument("--export-excel", type=str, default=None, help="Export completed rows to Excel (.xlsx)")
    parser.add_argument("--export-csv", type=str, default=None, help="Export completed rows to CSV (.csv)")
    parser.add_argument("--sample", type=int, default=None, help="Process a sample of N rows and export to Excel")
    parser.add_argument("--all", action="store_true", help="Process ALL remaining questions in dataset")
    args = parser.parse_args()

    engine = JeopardyTranslatorEngine(num_threads=args.threads)
    engine.import_csv_if_needed()

    if args.sample:
        print(f"=== Processing Sample of {args.sample} Rows ===", flush=True)
        engine.process_batch(limit=args.sample)
        engine.export_to_excel("JEOPARDY_GEORGIAN_SAMPLE.xlsx", limit=args.sample)
        engine.export_to_csv("JEOPARDY_GEORGIAN_SAMPLE.csv", limit=args.sample)
        return

    if args.all or args.limit:
        engine.process_batch(limit=args.limit, chunk_size=args.chunk_size)

    if args.export_excel:
        engine.export_to_excel(args.export_excel, limit=args.limit)

    if args.export_csv:
        engine.export_to_csv(args.export_csv, limit=args.limit)


if __name__ == "__main__":
    main()
